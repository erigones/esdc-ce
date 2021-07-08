from django.http import HttpResponse
from django.conf import settings
from django.utils import six

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import NamedStyle, PatternFill, Font, Color, Border, Side, Alignment
from openpyxl.styles.borders import BORDER_THIN


class Excel(object):

    wb = None

    # Default color definitions
    header_color = 'FF000000'
    header_bgcolor = 'FF98B4D4'
    odd_row = 'FFEBEDED'
    even_row = 'FFDEDCDC'

    def load_workbook(self, filename):
        self.wb = load_workbook(filename=filename, use_iterators=True)

    @staticmethod
    def set_cell_style(cell=None, color='FF000000', bgcolor='FFFFFFFF', font='Calibri', size=11, bold=False,
                       italic=False, underline='none', strike=False, border=None, border_style=BORDER_THIN,
                       border_bottom=None, border_bottom_style=None, horizontal='general', vertical='bottom',
                       number_format=None):
        if not border:
            border = 'FFB6B6B4'
        if not border_bottom:
            border_bottom = border
        if not border_bottom_style:
            border_bottom_style = border_style

        cell.font = Font(name=font, size=size, bold=bold, italic=italic, vertAlign=None, underline=underline,
                         strike=strike, color=color)
        cell.fill = PatternFill(patternType='solid', fgColor=Color(bgcolor))
        cell.border = Border(left=Side(border_style=border_style, color=Color(border)),
                             right=Side(border_style=border_style, color=Color(border)),
                             top=Side(border_style=border_style, color=Color(border)),
                             bottom=Side(border_style=border_bottom_style, color=Color(border_bottom)),
                             )
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=0, wrap_text=False,
                                   shrink_to_fit=False, indent=0)
        cell.number_format = number_format

        return cell

    @classmethod
    def update_cell(cls, sheet, ci, ri, value, color='FF000000', bgcolor='FFFFFFFF', font='Calibri', size=11,
                    bold=False, italic=False, underline='none', strike=False, border=None, border_style=BORDER_THIN,
                    horizontal='general', vertical='bottom', number_format=None, comment=None, author=None):
        if color:
            cls.set_cell_style(cell=sheet[ci + str(ri)], color=color, bgcolor=bgcolor, font=font, size=size, bold=bold,
                               italic=italic, underline=underline, strike=strike, border=border,
                               border_style=border_style, horizontal=horizontal, vertical=vertical,
                               number_format=number_format)
        if comment:
            if not author:
                author = settings.COMPANY_NAME
            sheet[ci + str(ri)].comment = Comment(comment, author)
        sheet[ci + str(ri)] = value

    def set_row_color(self, sheet, row, header, color='FF000000', bgcolor='FFFFFFFF'):
        for idx, item in enumerate(header):
            self.set_cell_style(cell=sheet[str(six.unichr(idx + ord('A'))) + str(row)], color=color, bgcolor=bgcolor)
        return row

    def clean_sheet(self, sheet, header, ignore_rows=0):
        """
        Clean all data in the sheet, by default leave header
        """
        for idx, column in enumerate(sheet.columns):
            if idx == len(header):
                break
            row = 0
            for cell in column:
                if row >= ignore_rows:
                    cell.value = None
                    self.set_cell_style(cell=cell)
                row += 1

    def offer_download(self, fname):
        response = HttpResponse(content_type='application/vms.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=%s' % fname

        self.wb.save(response)
        return response
