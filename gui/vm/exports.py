from __future__ import division
import random
from django.utils.six import text_type
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from gui.excel import Excel
from gui.vm.utils import get_vms, get_vm_define_disk, get_vm_define_nic, ImportExportBase
from api.vm.utils import get_subnets, get_images, get_templates, get_zpools, get_nodes
from vms.models import Vm


def generate_vms_file(request, hostnames):
    ee = Export(request)
    ee.generate_datacenter(request, hostnames=hostnames)

    return ee.offer_download('dc_%s_import.xlsx' % request.dc.name)


def generate_sample_file(request):
    ee = Export(request)
    ee.generate_datacenter(request, sample=True)

    return ee.offer_download('dc_%s_sample_import.xlsx' % request.dc.name)


class SampleTags(object):
    @staticmethod
    def names():
        tags = (('TagX',), ('TagY',), ('TagZ',), ('',), ('TagX', 'Client1'), ('TagY', 'TagZ'), ('Internal1', 'TagZ'))
        return random.choice(tags)


class SampleVm(object):
    request = None
    hostname = None
    alias = None
    tags = None
    node = None
    ostype = None
    vcpus = None
    ram = None

    def __init__(self, request, number, hostname='.cust.example.com'):
        self.request = request
        names = ('webserver', 'database', 'loadbalancer')

        self.alias = random.choice(names) + str(number)
        self.hostname = self.alias + hostname
        self.tags = SampleTags()
        self.node = ''
        self.vcpus = random.randint(1, 4)
        self.ram = random.choice((256, 512, 1024, 1536, 2048, 4096))

    # noinspection PyMethodMayBeStatic
    def get_ostype_display(self):
        return 'Linux VM'

    def generate_nics(self):
        nics = []
        nets = get_subnets(self.request)
        for i in range(0, random.randint(1, 2)):
            if len(nets) > 0:
                random_net = random.randint(0, len(nets) - 1)
                net = nets[random_net].name
            else:
                net = ''

            nics.append({
                'net': net,
                'ip': ''
            })

        return nics

    def generate_disks(self):
        disks = []
        images = get_images(self.request, ostype=1)
        zpools = get_zpools(self.request).distinct('zpool')
        for i in range(0, random.randint(1, 2)):
            if len(zpools) > 0:
                random_zpool = random.randint(0, len(zpools) - 1)
                zpool = zpools[random_zpool].zpool
            else:
                zpool = ''

            disk = {
                'zpool': zpool,
                'size': random.choice((10245, 15365, 20485, 25605)),
                'image': '',
            }

            if i == 0 and len(images) > 0:
                random_image = random.randint(0, len(images) - 1)
                disk['image'] = images[random_image].name

            disks.append(disk)

        return disks


class Export(ImportExportBase, Excel):

    request = None

    def __init__(self, request):
        super(Export, self).__init__()

        self.request = request
        # Create new Workbook with appropriate sheets
        self.wb = Workbook()

        self.sheet_dc = self.wb.active
        self.sheet_dc.title = self.sheet_dc_name
        self.sheet_dc.freeze_panes = self.sheet_dc.cell('B3')
        self.sheet_dc_generate_header()

        self.sheet_data = self.wb.create_sheet(title=self.sheet_data_name)
        self.sheet_data.freeze_panes = self.sheet_data.cell('A2')
        self.sheet_data_generate_content(set_validation=False)

    def sheet_dc_generate_header(self):
        self.set_row_color(self.sheet_dc, 1, self.get_header(), color=self.header_color,
                           bgcolor=self.header_bgcolor)
        self.update_cell(self.sheet_dc, 'A', 1, 'Total', bold=True, size=12, color=self.header_color,
                         bgcolor=self.header_bgcolor, horizontal='center')

        for idx, typex in enumerate(self.get_file_header()):
            comment = None
            if typex in ('RAM', 'HDD Size'):
                comment = 'Size in GB'
            if typex == 'Node':
                comment = 'When node is not selected, auto-select is applied during the import'
            if typex == 'Tags':
                comment = 'Multiple tags are supported, separate them with comma'
            self.update_cell(self.sheet_dc, self.HEADER_MAP[idx][0], 2, typex, color=self.header_color,
                             bgcolor=self.header_bgcolor, horizontal='center', comment=comment)

            if typex in ('vCPU', 'RAM', 'HDD Size'):
                self.update_cell(self.sheet_dc, self.HEADER_MAP[idx][0], 1,
                                 '=SUM(' + self.HEADER_MAP[idx][0] + '3:' + self.HEADER_MAP[idx][0] + '1000)',
                                 bold=True, color=self.header_color, bgcolor=self.header_bgcolor)

    def generate_datacenter(self, request, sample=False, hostnames=None):
        row = 3
        row_color = self.even_row

        if sample:
            vms = []
            for i in range(0, 3):
                vms.append(SampleVm(request, number=i + 1))
        else:
            vms = get_vms(request).filter(hostname__in=hostnames)

        for vm in vms:
            # Decide what color row should have
            if row_color == self.even_row:
                row_color = self.odd_row
            else:
                row_color = self.even_row

            nic_row = disk_row = row
            str_row = str(row)
            self.set_row_color(self.sheet_dc, str_row, self.get_header(), bgcolor=row_color)

            # Store datacenter into sheet
            self.sheet_dc[self.get_column_index('Hostname') + str_row] = vm.hostname
            self.sheet_dc[self.get_column_index('Alias') + str_row] = vm.alias
            self.sheet_dc[self.get_column_index('Tags') + str_row] = ','.join([t for t in vm.tags.names()])
            if vm.node:
                if request.user.is_admin(request) and vm.node.hostname:
                    self.sheet_dc[self.get_column_index('Node') + str_row] = vm.node_hostname
                else:
                    self.update_cell(self.sheet_dc, self.get_column_index('Node'), str_row, '',
                                     bgcolor=vm.node.color.replace('#', 'FF'))

            self.sheet_dc[self.get_column_index('OS Type') + str_row] = vm.get_ostype_display()
            self.sheet_dc[self.get_column_index('vCPU') + str_row] = vm.vcpus
            self.sheet_dc[self.get_column_index('RAM') + str_row] = self._convert_to_gb(vm.ram)

            if sample:
                vm_nics = vm.generate_nics()
            else:
                vm_nics = get_vm_define_nic(request, vm)

            for vm_nic in vm_nics:
                if row < nic_row:
                    self.set_row_color(self.sheet_dc, nic_row, self.get_header(), bgcolor=row_color)
                self.sheet_dc[self.get_column_index('Network') + str(nic_row)] = vm_nic['net']
                self.sheet_dc[self.get_column_index('IP') + str(nic_row)] = vm_nic['ip']
                nic_row += 1

            if sample:
                vm_disks = vm.generate_disks()
            else:
                vm_disks = get_vm_define_disk(request, vm)

            for vm_disk in vm_disks:
                if row < disk_row:
                    self.set_row_color(self.sheet_dc, disk_row, self.get_header(), bgcolor=row_color)

                self.sheet_dc[self.get_column_index('Storage') + str(disk_row)] = vm_disk['zpool']
                self.sheet_dc[self.get_column_index('HDD Size') + str(disk_row)] = self._convert_to_gb(vm_disk['size'])
                if vm_disk['image'] is not None:
                    self.sheet_dc[self.get_column_index('Image') + str(disk_row)] = vm_disk['image']
                disk_row += 1

            if disk_row > nic_row:
                row = disk_row
            else:
                row = nic_row

    @staticmethod
    def _convert_to_gb(num):
        return round(num / 1024, 1)

    #
    # Functions required for Data sheet generation
    #

    def add_validation(self, column, no_rows, validation_column):
        """
        Function that creates excel validation, based on columns in Data sheet and applies it in Datacenter sheet
        """
        # Define validation column with number of rows
        dv_str = '=Data!$' + column + '$2:$' + column + '$' + str(no_rows + 25)
        dv = DataValidation('list', formula1=dv_str, allow_blank=True)
        self.sheet_dc.add_data_validation(dv)
        # Apply newly created validation in Datacenter sheet to validate against generated values
        dv.ranges.append(str(validation_column) + '3:' + str(validation_column) + '1048576')

    def generate_networks(self, column_index, row_index=1, label='Network'):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for row_index, network in enumerate(get_subnets(self.request), start=row_index + 1):
            self.update_cell(self.sheet_data, column_index, row_index, network.name)

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def generate_images(self, column_index, row_index=1, label='Image'):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for row_index, image in enumerate(get_images(self.request), start=row_index + 1):
            self.update_cell(self.sheet_data, column_index, row_index, image.name)

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def generate_ostypes(self, column_index, row_index=1, label='OS Type'):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for idx, os in Vm.OSTYPE:
            row_index += 1
            self.update_cell(self.sheet_data, column_index, row_index, text_type(os))

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def generate_templates(self, column_index, row_index=1, label='Template'):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for row_index, template in enumerate(get_templates(self.request), start=row_index + 1):
            self.update_cell(self.sheet_data, column_index, row_index, template.name)

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def generate_storages(self, column_index, row_index=1, label='Storage'):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for row_index, ns in enumerate(get_zpools(self.request).distinct('zpool'), start=row_index + 1):
            self.update_cell(self.sheet_data, column_index, row_index, ns.zpool)

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def generate_nodes(self, column_index, row_index=1, label='Node', color=False):
        self.update_cell(self.sheet_data, column_index, row_index, label, color=self.header_color,
                         bgcolor=self.header_bgcolor)

        for row_index, node in enumerate(get_nodes(self.request), start=row_index + 1):
            if color:
                self.update_cell(self.sheet_data, column_index, row_index, None, color=node.color.replace('#', 'FF'))
            else:
                self.update_cell(self.sheet_data, column_index, row_index, node.hostname)

        self.add_validation(column_index, row_index, self.get_column_index(label))

    def sheet_data_generate_content(self, set_validation=False):
        # self.clean_sheet(self.sheet_data, 1)
        if self.request.user.is_admin(self.request):
            self.generate_nodes('A')
        else:
            self.generate_nodes('A', True)
        self.generate_ostypes('B')
        self.generate_networks('C')
        self.generate_storages('D')
        self.generate_images('E')
        # self.generate_backup('F') # Backup will be part of 2.1 or later....
        # self.generate_templates('G') # We decided not to export templates

        if set_validation:
            # Set validation for numbers
            ci = self.get_column_index('RAM')
            ram_validation = DataValidation(type='whole', operator='between', formula1='32', formula2='524288')
            self.sheet_dc.add_data_validation(ram_validation)
            ram_validation.ranges.append(ci + '2:' + ci + '1048576')

            ci = self.get_column_index('vCPU')
            cpu_validation = DataValidation(type='whole', operator='between', formula1='1', formula2='64')
            self.sheet_dc.add_data_validation(cpu_validation)
            cpu_validation.ranges.append(ci + '2:' + ci + '1048576')

            ci = self.get_column_index('HDD Size')
            hdd_validation = DataValidation(type='whole', operator='between', formula1='10240', formula2='268435456')
            self.sheet_dc.add_data_validation(hdd_validation)
            hdd_validation.ranges.append(ci + '2:' + ci + '1048576')
